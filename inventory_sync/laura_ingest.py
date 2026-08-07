"""Laura ingest core: parse the supplier xlsx and create net-new draft products.

Wires grouping (laura_upload) + mapping (laura_mapping) + the store create seams
+ store_products pending rows. See tests/test_laura_ingest.py.

Ingest is skip-dominant: a typical upload has ~2025 rows, almost all already on
the store; only a handful survive as new. New SKUs are assumed to be new products
(a title collision with an existing product is flagged for manual review, not
duplicated).
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field, replace
from decimal import Decimal, InvalidOperation

import openpyxl

from inventory_sync import review_reasons
from inventory_sync.domain import SKU
from inventory_sync.laura_mapping import (
    CATEGORY_COLLECTION_ID,
    VENDOR,
    subcategory_collection,
    to_product_draft,
)
from inventory_sync.missing_source import reconcile_missing_at_source
from inventory_sync.pricing import resolve_target
from inventory_sync.pricing_sync import reconcile_prices

# Skip missing-at-source flagging if the uploaded file covers less than this
# fraction of existing Laura products — it's probably a partial/wrong upload.
_MISSING_COVERAGE_MIN = 0.75
from inventory_sync.laura_upload import LauraRow, group_products
from inventory_sync.persistence.store_product_store import NewStoreProduct


@dataclass
class IngestSummary:
    created: int = 0
    skipped_existing: int = 0
    flagged_review: int = 0
    flagged_missing: int = 0    # Laura products gone from the file entirely (missing-at-source)
    prices_updated: int = 0     # existing products repriced from the new Excel
    prices_would_update: int = 0
    prices_blocked: int = 0
    errors: int = 0            # products that failed to create/archive (isolated, batch continues)
    would_create: int = 0      # dry-run: products that would be created
    archived: int = 0          # discontinued ("אזל") products taken down from the site
    would_archive: int = 0     # dry-run: products that would be taken down
    dry_run: bool = False
    created_skus: list[str] = field(default_factory=list)


_EXACT_HEADERS = {
    "מקט": "sku",
    "ברקוד": "barcode",
    "תיאור פריט": "description",
    "תאור משפחה": "family",
    "טקסט": "text",
    "מחיר מומלץ": "recommended_price",
}

# Columns whose ABSENCE breaks the ingest — most importantly `availability`
# ("מלאי"): without it we can't tell what's sold out ("אזל"), so the take-down
# filter silently no-ops and the whole catalog (incl. discontinued items) gets
# onboarded as drafts. If any of these is missing we refuse the file rather than
# guess. Value = a human hint at the expected header.
_REQUIRED_COLUMNS = {
    "sku": "מקט",
    "description": "תיאור פריט",
    "family": "תאור משפחה",
    "availability": "מלאי / מלאי זמין",
}


class LauraFileError(ValueError):
    """The uploaded Laura xlsx doesn't have the expected structure (missing columns)."""


def _s(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _dec(value) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def parse_laura_xlsx(data: bytes) -> list[LauraRow]:
    """Parse supplier xlsx bytes into LauraRows via header-based column mapping."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    col: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        hs = str(h).strip()
        if hs in _EXACT_HEADERS:
            col[_EXACT_HEADERS[hs]] = i
        elif "קישור" in hs or "link" in hs.lower():
            col["image_url"] = i
        elif "מלאי" in hs or "זמין" in hs:  # stock/availability: "מלאי", "מלאי זמין", "מלרי זמין"
            col["availability"] = i

    # Fail loudly on a wrong-structure file instead of silently onboarding it.
    missing = [f"{field} ({hint})" for field, hint in _REQUIRED_COLUMNS.items() if field not in col]
    if missing:
        seen = [str(h).strip() for h in header if h not in (None, "")]
        raise LauraFileError(
            "Laura file is missing required column(s): " + ", ".join(missing)
            + f". Headers found: {seen}"
        )

    def cell(row, field_name):
        i = col.get(field_name)
        return row[i] if i is not None and i < len(row) else None

    out: list[LauraRow] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        sku = _s(cell(row, "sku"))
        if not sku:
            continue
        out.append(LauraRow(
            sku=sku,
            description=_s(cell(row, "description")) or "",
            family=_s(cell(row, "family")) or "",
            barcode=_s(cell(row, "barcode")),
            text=_s(cell(row, "text")),
            image_url=_s(cell(row, "image_url")),
            recommended_price=_dec(cell(row, "recommended_price")),
            availability=_s(cell(row, "availability")),
        ))
    return out


def _create_and_record(store, product_store, customer_id, group, draft, sub_name, review_reason, logger) -> None:
    """Create one product, attach collections, record the pending row.

    Product is created FIRST so a failed create leaves no orphan collection.
    """
    created = store.create_product(draft)
    sub_ref = store.ensure_collection(sub_name) if sub_name else None
    store.add_to_collection(created.store_product_id, CATEGORY_COLLECTION_ID)
    if sub_ref is not None:
        store.add_to_collection(created.store_product_id, sub_ref.id)
    is_new_collection = bool(sub_ref and sub_ref.created)
    product_store.write_pending(customer_id, [
        NewStoreProduct(
            sku=str(v.sku),
            store_product_id=created.store_product_id,
            title=group.title,
            vendor=draft.vendor,
            is_new_collection=is_new_collection,
            needs_review=review_reason is not None,
            needs_review_reason=review_reason,
        )
        for v in group.variants
    ])
    logger.info("ingest_created", title=group.title, store_product_id=created.store_product_id,
                variants=len(group.variants), is_new_collection=is_new_collection,
                needs_review_reason=review_reason)


def ingest_products(rows, store, product_store, customer_id: str, logger, dry_run: bool = False,
                    sync_prices: bool = False, price_dry_run: bool = True) -> IngestSummary:
    """Group rows, take down discontinued items, create new products as drafts.

    `מלאי זמין` = "אזל" (discontinued) rows are never uploaded, and are archived
    if already on the site. Note: unpublish acts at product level, so a product
    whose SKU is discontinued is archived as a whole.
    """
    summary = IngestSummary(dry_run=dry_run)
    existing = store.list_products()
    existing_by_sku = {str(p.sku): p for p in existing}
    existing_skus = set(existing_by_sku)
    existing_titles = {p.title for p in existing if p.title}

    # --- Takedown pass: discontinued ("אזל") items removed from the site (archive) ---
    for sku in sorted({r.sku for r in rows if r.discontinued}):
        prod = existing_by_sku.get(sku)
        if prod is None or not prod.published:
            continue  # not on site, or already down — nothing to take down
        if dry_run:
            summary.would_archive += 1
            logger.info("ingest_would_archive", sku=sku)
            continue
        try:
            store.unpublish(SKU(sku))
            summary.archived += 1
            logger.info("ingest_archived_discontinued", sku=sku)
        except Exception as e:
            summary.errors += 1
            logger.error("ingest_archive_failed", sku=sku, error=str(e)[:200])

    # --- Create pass: only active (non-discontinued) rows are eligible to upload ---
    active_rows = [r for r in rows if not r.discontinued]
    for group in group_products(active_rows):
        group_skus = {str(v.sku) for v in group.variants}

        if group_skus & existing_skus:
            summary.skipped_existing += 1
            logger.info("ingest_skip_existing", title=group.title, skus=sorted(group_skus))
            continue

        if group.title in existing_titles:
            # New SKU whose title matches an existing product — likely a new size of
            # an existing product (add-variant, not create). Flag, don't duplicate.
            summary.flagged_review += 1
            logger.warning("ingest_title_collision", title=group.title, skus=sorted(group_skus))
            continue

        sub_name = subcategory_collection(group.family)
        review_reason = review_reasons.join(
            review_reasons.SUPPLIER_FLAG if group.needs_review else None,
            review_reasons.NO_COLLECTION if sub_name is None else None,
            review_reasons.NO_IMAGE if not group.image_urls else None,
            review_reasons.NO_BODY if not group.body_text else None,
        )

        if dry_run:
            summary.would_create += 1
            logger.info("ingest_would_create", title=group.title,
                        variants=len(group.variants), needs_review_reason=review_reason)
            continue

        draft = to_product_draft(group)
        try:
            _create_and_record(store, product_store, customer_id, group, draft, sub_name, review_reason, logger)
        except Exception as first_err:
            # A bad image URL is the common failure (Shopify 422). Salvage the
            # product by retrying once WITHOUT images, flagged for review.
            if draft.image_urls:
                try:
                    _create_and_record(store, product_store, customer_id, group,
                                       replace(draft, image_urls=()), sub_name,
                                       review_reasons.IMAGE_REJECTED, logger)
                    summary.created += 1
                    summary.created_skus.extend(sorted(group_skus))
                    logger.warning("ingest_created_without_image", title=group.title,
                                   error=str(first_err)[:200])
                    continue
                except Exception as retry_err:
                    first_err = retry_err
            # Isolated failure — record and move on; never abort the batch.
            logger.error("ingest_create_failed", title=group.title,
                         skus=sorted(group_skus), error=str(first_err)[:200])
            summary.errors += 1
            continue
        summary.created += 1
        summary.created_skus.extend(sorted(group_skus))

    # --- missing-at-source: Laura store products absent from THIS file entirely ---
    # (Not the "אזל" takedown above — that's for in-file discontinued rows. This
    # catches products that dropped off the file completely.) GUARDED against a
    # partial/wrong upload: if the file covers <75% of existing Laura products it
    # is probably incomplete, so we skip flagging rather than flag hundreds.
    file_skus = {str(r.sku) for r in rows if r.sku}
    laura_skus = {str(p.sku) for p in existing if p.vendor == VENDOR}
    if laura_skus:
        coverage = len(laura_skus & file_skus) / len(laura_skus)
        if coverage < _MISSING_COVERAGE_MIN:
            logger.warning("ingest_missing_check_skipped_partial_file",
                           coverage=round(coverage, 3), laura_products=len(laura_skus),
                           file_skus=len(file_skus), threshold=_MISSING_COVERAGE_MIN)
        else:
            summary.flagged_missing = reconcile_missing_at_source(
                product_store, existing, file_skus, (VENDOR,),
                customer_id, logger, dry_run=dry_run)

    # --- price sync (Excel upload only): reprice existing Laura products from the
    # new file. Target = the Excel price (the base×1.77 shelf price). Baseline
    # unless price is live; write-avoidance + >60% guard + Axiom audit apply. ---
    if sync_prices:
        targets = {str(r.sku): resolve_target(r.recommended_price, None)
                   for r in rows if not r.discontinued and r.recommended_price is not None}
        psum = reconcile_prices(store, existing, targets, logger,
                                dry_run=dry_run or price_dry_run)
        summary.prices_updated = psum.updated
        summary.prices_would_update = psum.would_update
        summary.prices_blocked = psum.blocked

    logger.info("ingest_summary", customer_id=customer_id, created=summary.created,
                skipped_existing=summary.skipped_existing, flagged_review=summary.flagged_review,
                errors=summary.errors, would_create=summary.would_create,
                archived=summary.archived, would_archive=summary.would_archive,
                flagged_missing=summary.flagged_missing, prices_updated=summary.prices_updated,
                prices_would_update=summary.prices_would_update, dry_run=dry_run)
    return summary
